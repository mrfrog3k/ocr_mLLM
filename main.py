import requests
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from PIL import Image, ImageTk
import os
import datetime
import base64
from openpyxl import Workbook, load_workbook

# OCR API call
def perform_ocr(image_url=None, image_base64=None):
    payload = {}
    if image_url:
        payload["image_url"] = image_url
    elif image_base64:
        payload["image_base64"] = image_base64
    else:
        return "Không có ảnh để gửi.", None

    try:
        response = requests.post(
            "https://8620-34-169-92-156.ngrok-free.app/ocr",
            json=payload
        )
        if response.status_code == 200:
            return response.json().get("response_message"), response.json().get("raw_output")
        else:
            return f"Lỗi {response.status_code}: {response.text}", None
    except Exception as e:
        return f"Exception: {str(e)}", None

# Mã hóa ảnh sang base64
def encode_image_base64(filepath):
    with open(filepath, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

# Phân tích dữ liệu dạng bảng
def parse_ocr_table(text):
    lines = [line.strip() for line in text.strip().split("\n") if line.strip()]
    headers = []
    data_rows = []

    for line in lines:
        if line.startswith("|") and line.endswith("|"):
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if not headers:
                headers = cells
            else:
                while len(cells) < len(headers):
                    cells.append("")
                while len(cells) > len(headers):
                    headers.append(f"Cột {len(headers)+1}")
                data_rows.append(cells)

    return headers, data_rows

# Gửi URL
def on_submit():
    image_url = url_entry.get().strip()
    if not image_url:
        messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập đường dẫn ảnh.")
        return

    loading_label.config(text="Đang xử lý...", fg="blue")
    root.update()

    result, raw_output = perform_ocr(image_url=image_url)
    loading_label.config(text="")
    handle_result(result, raw_output)

# Chọn ảnh từ thiết bị
def on_select_file():
    file_path = filedialog.askopenfilename(
        title="Chọn ảnh hóa đơn",
        filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp")]
    )
    if file_path:
        display_image_preview(file_path)
        loading_label.config(text="Đang xử lý...", fg="blue")
        root.update()
        image_base64 = encode_image_base64(file_path)
        result, raw_output = perform_ocr(image_base64=image_base64)
        loading_label.config(text="")
        handle_result(result, raw_output)

# Hiển thị ảnh xem trước
def display_image_preview(filepath):
    try:
        img = Image.open(filepath)
        img.thumbnail((200, 200))  # Giới hạn kích thước hiển thị
        photo = ImageTk.PhotoImage(img)
        preview_label.config(image=photo)
        preview_label.image = photo  # Giữ tham chiếu để tránh bị garbage collected
    except Exception as e:
        messagebox.showerror("Lỗi hiển thị", f"Không thể hiển thị ảnh: {e}")

# Hiển thị dữ liệu trả về
def handle_result(result, raw_output):
    if not result:
        messagebox.showerror("Lỗi", "Không kết nối được với Server hoặc có lỗi xảy ra.")
        return

    # Hiển thị response_message (kết quả dạng bảng) trong raw_output_text
    raw_output_text.delete("1.0", tk.END)
    raw_output_text.insert(tk.END, result if result else "Không có dữ liệu response.")

    if "|" not in result:
        messagebox.showinfo("Thông tin", "Không tìm thấy dữ liệu hợp lệ. Lỗi kết nối đến server")
    else:
        headers, rows = parse_ocr_table(result)
        if headers and rows:
            create_tree(headers)
            for row in rows:
                tree.insert("", tk.END, values=row)
            global current_headers
            current_headers = headers
        else:
            messagebox.showinfo("Không có dữ liệu", "Không tìm thấy bảng hợp lệ để hiển thị ở bảng.")

# Tạo bảng Treeview
def create_tree(headers):
    global tree, tree_frame
    if tree:
        tree.destroy()

    tree = ttk.Treeview(tree_frame, columns=headers, show="headings")

    for col in headers:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor=tk.CENTER)

    tree.pack(fill=tk.BOTH, expand=True)
    tree.bind("<Double-1>", on_double_click)

# Sửa dữ liệu trong bảng
def on_double_click(event):
    global edit_entry

    if edit_entry:
        edit_entry.destroy()
        edit_entry = None

    region = tree.identify("region", event.x, event.y)
    if region != "cell":
        return

    row_id = tree.identify_row(event.y)
    column_id = tree.identify_column(event.x)
    if not row_id or not column_id:
        return

    x, y, width, height = tree.bbox(row_id, column_id)
    column_index = int(column_id[1:]) - 1
    current_value = tree.item(row_id)["values"][column_index]

    edit_entry = tk.Entry(tree)
    edit_entry.place(x=x, y=y, width=width, height=height)
    edit_entry.insert(0, current_value)
    edit_entry.focus()

    def save_edit(event):
        new_value = edit_entry.get()
        values = list(tree.item(row_id)["values"])
        values[column_index] = new_value
        tree.item(row_id, values=values)
        edit_entry.destroy()

    edit_entry.bind("<Return>", save_edit)
    edit_entry.bind("<FocusOut>", lambda e: edit_entry.destroy())

# Xuất ra Excel
def export_to_excel():
    if not current_headers or not tree.get_children():
        messagebox.showwarning("Không có dữ liệu", "Không có bảng nào để xuất.")
        return

    folder = "data_result"
    os.makedirs(folder, exist_ok=True)
    filename = os.path.join(folder, "ocr_data.xlsx")

    file_exists = os.path.exists(filename)

    if file_exists:
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Result"
        ws.append(current_headers)  # Ghi header nếu tạo file mới

    # Ghi dữ liệu từ bảng
    for row_id in tree.get_children():
        row = tree.item(row_id)["values"]
        ws.append(row)

    wb.save(filename)
    messagebox.showinfo("Xuất thành công", f"Dữ liệu đã được thêm vào file: {filename}")

# Giao diện chính
root = tk.Tk()
root.title("OCR từ ảnh hóa đơn")
root.geometry("1000x700")

# Sử dụng grid layout cho tổ chức tốt hơn
main_frame = ttk.LabelFrame(root, text="Nhập liệu và Điều khiển")
main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

tk.Label(main_frame, text="Nhập đường dẫn ảnh hóa đơn:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
url_entry = ttk.Entry(main_frame, width=80)
url_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
url_entry.bind("<Return>", lambda event: on_submit())  # Gửi khi nhấn Enter

submit_button = ttk.Button(main_frame, text="Gửi yêu cầu OCR (từ URL)", command=on_submit)
submit_button.grid(row=1, column=0, columnspan=2, pady=5)

select_button = ttk.Button(main_frame, text="Chọn ảnh từ thiết bị", command=on_select_file)
select_button.grid(row=2, column=0, columnspan=2, pady=5)

# Khu vực xem trước ảnh
preview_frame = ttk.LabelFrame(root, text="Xem trước ảnh")
preview_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
preview_label = tk.Label(preview_frame)
preview_label.pack(padx=5, pady=5)

# Khu vực hiển thị raw output (thực tế sẽ hiển thị response_message)
raw_output_frame = ttk.LabelFrame(root, text="Kết quả OCR")
raw_output_frame.grid(row=0, column=1, rowspan=2, padx=10, pady=10, sticky="nsew")
raw_output_text = scrolledtext.ScrolledText(raw_output_frame, width=50, height=15)
raw_output_text.pack(fill="both", expand=True, padx=5, pady=5)

# Khu vực hiển thị bảng
table_frame = ttk.LabelFrame(root, text="Dữ liệu OCR dạng bảng")
table_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
tree_frame = tk.Frame(table_frame)
tree_frame.pack(fill=tk.BOTH, expand=True)
tree = None

export_button = ttk.Button(table_frame, text="Xuất Excel", command=export_to_excel)
export_button.pack(pady=10)

loading_label = tk.Label(root, text="", font=("Arial", 10))
loading_label.grid(row=3, column=0, columnspan=2, pady=5)

edit_entry = None
current_headers = []

# Cấu hình grid để các frame có thể thay đổi kích thước
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)

root.mainloop()